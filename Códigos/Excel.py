from openpyxl import Workbook  # W mayúscula
wb = Workbook()  # W mayúscula
ws0 = wb.active  # No se está claro que es lo que hace
# Se coloca al final de las hojas de Excel
ws1 = wb.create_sheet("My sheet hello")
# Al definir la posición, puedo colocar la hoja donde quiera
ws2 = wb.create_sheet("My sheet 0", 0)
# cambiar el nombre de las hojas de cálculo.
ws2.title = "My edited name sheet"
print(wb.sheetnames)  # imprimir los nombres de las hojas
ws0.sheet_properties.tabColor = "1072BA"  # Color de la pestaña
# esta variable aparentemente está asignando el contenido de la hoja que
ws3 = wb["My sheet hello"]
# se llama como he descrito. Puedes hacer búsquedas en libro de cálculo de Excel usando
# el nombre de sus pestañas como índice.
print(ws3)  # imprime la hoja de cálculo
for sheet in wb:  # los libros de cálculo son interables
    print(sheet.title)  # imprimiendo título de hoja uno a la vez
source = wb.active  # ...
target = wb.copy_worksheet(source)  # Copiando un libro de cálculo.
# Jugando con datos
c = ws0["A4"]  # Asignando contenido de celda A4 de la hoja ws0 a variable c
print(c)  # imprimiendo celda
ws0["A4"] = 4  # asignando un nuevo valor a la celda A4 de la hoja ws0
c = ws0["A4"]  # Reasignando el valor de c
print(c)  # imprime la celda
# También se puede acceder a las celdas por este medio
d = ws0.cell(row=1, column=1, value=2)
print(d)  # imprimiendo la celda
e = ws0.cell(5, 6, 10)  # fila 5, columna 6 y valor 10
print(e)  # imprimiendo la celda
cell_range = ws0["A1":"C5"]  # Aisgnar un rango de Excel a una variable
# Nota: Rangos en diagonal como el descrito arriba representan todas las filas y columnas
# contenidas entre los extremos colocados.
print(cell_range)  # Imprimiendo rangos.
# Usar esto para aumentar la cantidad de elementos en C
cell_range = ws0["A1":"C8"]
# en la línea siguiente
colC = ws0["C"]  # Asignar columna C
# Imprimir columna (la cantidad de elementos impresos depende las celdas que han sido
print(colC)
# Creadas en memoria).
col_range = ws0["C:D"]
# Imprimir columnas. Se crea un mínimo de 5 elementos en la columna D ya que
print(col_range)
# no se habían creado elementos en D hasta entonces
# Ahora sí se define la cantidad de elementos en D (8 elementos).
cell_range = ws0["A1":"D8"]
print(cell_range)  # En este rango se verán los 8 elementos de D.
row10 = ws0[10]  # Creación rango basados en los elementos de  la fila 10
# Ésta fila no había sido usada hasta ahora, por lo que se crea una canidad
# de elementos por defecto en dicha  fila.
# Esto aumenta la cantidad de elementos creados en la línea siguiente
cell_range = ws0["A10":"I10"]
row10 = ws0[10]  # Creación rango basados en los elementos de  la fila 10
print(row10)  # imprimir rango de fila.
# rango entre los elementos de la fila 5 hasta la fila 10
row_range = ws0["5":"10"]
print(row_range)  # imprimir rango de filas
for fila in ws0.iter_rows(min_row=1, max_col=4, max_row=5):
    for celda in fila:
        print(celda)
# el primer for es un recorrido por medio de la función iter_rows
# que vuelve las filas un objeto iterable desde una fila mínima
# hasta una máxima delimitano el tamaño de las filas por medio
# de una columna de frontera max_col. El segundo for usa a
# fila como objeto iterable al imprimir uno a uno las celdas
#contenidas en fila
print(fila) #Acá sólo se ven los úlrimos elementos de las filas
# pero se tiene acceso por medio de ciclos a cada una e las 
# celdas contenidas en las filas.
for col in ws0.iter_cols(min_row=1, max_col=3, max_row=2):
        for celda in col:
                print(celda)